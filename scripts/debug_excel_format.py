import openpyxl
import sys

def debug_excel(file_path):
    print(f"--- Debugging: {file_path} ---")
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        print(f"Sheet name: {sheet.title}")
        print(f"Max Column: {sheet.max_column}")
        
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=15), 1):
            row_data = []
            for cell in row:
                row_data.append(f"[{cell.column_letter}]{str(cell.value).strip() if cell.value is not None else 'EMPTY'}")
            print(f"Row {i}: {row_data}")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        debug_excel(sys.argv[1])
    else:
        print("Please provide a file path.")
